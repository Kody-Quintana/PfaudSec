from openpyxl.utils import column_index_from_string, get_column_letter
from pylatexenc.latexencode import utf8tolatex
from PyQt5 import QtWidgets, QtCore, QtGui
from openpyxl import load_workbook
from collections import Counter
#import Cryptodome (imported in a try statement below)
import dateutil.relativedelta
import configparser
import qdarkstyle
import traceback
import functools
import itertools
import datetime
import tempfile
import pathlib
import shutil
import time
import sys
import io
import os

import ini_storage # Stores example config
import font_load   # Decrypt fonts
import ui_ini_edit    # Basic text editor for ini files
import ui_date_set    # QDialog to change date
import ui_log      # Log window

#pgfplots tex file stored as list in a python file
from tex_storage import line_graph_tex, bar_graph_tex, percent_line_graph_tex, latex_table, cost_avg_graph
from custom_page_tex import custom_page_tex

def intersperse(lst, item, begin=None, end=None):
    result = [item] * (len(lst) * 2 - 1)
    result[0::2] = lst
    if end != None:
        result.append(end)
    if begin != None:
        result.insert(0, begin)
    return result

if getattr(sys, 'frozen', False):
    """Chdir to exe if frozen with pyinstaller"""
    os.chdir(os.path.dirname(sys.executable))

def fy_date(date):
    """Returns strftime('%B %Y') style date but with fiscal year"""
    month = date.strftime('%B')
    if date.month in (9, 10, 11, 12):
        year = str(int(date.strftime('%Y')) + 1)
    else:
        year = date.strftime('%Y')
    return month + ', FY' + year

def fy_year(date):
    """Returns strftime('%Y') style year but with fiscal year"""
    if date.month in (9, 10, 11, 12):
        year = str(int(date.strftime('%Y')) + 1)
    else:
        year = date.strftime('%Y')
    return 'FY' + year

def folder_check(folder):
    """Convenience function to make sure folder exists and define variable in one line"""
    pathlib.Path(folder).mkdir(parents=True, exist_ok=True)
    return folder #so you can just do: folder_var = folder_check('string_path')

def filename_noext(filename):
    """Returns filename with no extension"""
    return filename.split('/')[len(filename.split('/')) - 1].rsplit(".", 1)[0]

def except_box(excType, excValue, tracebackobj):
    """Exceptions from sys.excepthook displayed in a QMessageBox and saved to logfile"""
    logFile = "PfaudSec_spreadsheet_crash.log"
    notice = 'An unhandled exception occured. Please report the problem via email to Quintana.Kody@gmail.com'\
            + '\n\nPlease attach the log file PfaudSec_spreadsheet_crash.log from the PfaudSec spreadsheet folder to the email.\n'
    versionInfo = "0.0.0"
    timeString = time.strftime("%Y-%m-%d, %H:%M:%S")

    tbinfofile = io.StringIO()
    traceback.print_tb(tracebackobj, None, tbinfofile)
    tbinfofile.seek(0)
    tbinfo = tbinfofile.read()
    errmsg = '%s: \n%s' % (str(excType), str(excValue))
    sep = '\n'
    sections = [sep, timeString, sep, errmsg, sep, tbinfo]
    msg = ''.join(sections)
    try:
        f = open(logFile, "w")
        f.write(msg)
        f.write(versionInfo)
        f.close()
    except IOError:
        pass
    errorbox = QtWidgets.QMessageBox()
    errorbox.setText(str(notice)+str(msg)+str(versionInfo))
    errorbox.exec_()

class EditConfig(QtWidgets.QDialog, ui_ini_edit.Ui_Dialog):
    """Basic text editor for document ini files

    will open with an example config for any new documents without an ini file"""
    def __init__(self, mode, file_to_save):
        super(EditConfig, self).__init__()
        self.setupUi(self)
        self.setModal(True)
        font = QtGui.QFontDatabase.systemFont(QtGui.QFontDatabase.FixedFont)
        font.setPointSize(14)
        self.textEdit.setFont(font)
        self.file_to_save = file_to_save
        self.buttonBox.button(QtWidgets.QDialogButtonBox.Save).clicked.connect(self.save_file)
        self.setWindowTitle(self.file_to_save)

        def is_not_empty(item):
            try:
                if os.path.getsize(item) > 0:
                    return True
                else:
                    return False
            except OSError:
                return False


        if is_not_empty(file_to_save):
            with open(file_to_save, 'r') as existing_file:
                self.textEdit.setText(existing_file.read())
        else:
            if mode == 'document':
                self.textEdit.setText('; "'
                        + self.file_to_save
                        + '"'
                        + ini_storage.ini_document)

    def save_file(self):
        """Save file and rebuild menu"""
        with open(self.file_to_save, 'w') as ini_file:
            ini_file.write(self.textEdit.toPlainText())
        trayIcon.rebuild_menu()

class Grapher(object):
    """Additional functions for openpyxl, LaTeX pgfplots writers, and a config file generator"""

    def __init__(self, work_dir, work_file, doc_name):
        """sets attributes based on config file, or prompts for a config file"""
        self.ready = False
        self.doc_name = filename_noext(doc_name)
        self.work_dir = work_dir
        self.work_file = work_file
        #self.months = months


        #Config file is just name of file with ini extension
        self.config_file = 'resource/' + filename_noext(self.work_file) + '.ini'
        self.config = configparser.ConfigParser()
        def try_config_read():
            try:
                with open(self.config_file) as f:
                    self.config.read_file(f)
            except IOError:
                edit_config = EditConfig('document', self.config_file)
                print(self.config_file)
                edit_config.exec_() #Block until closed
                try_config_read() #keep trying until the file exists and is read
        try_config_read()

        self.months = int(self.config.get('document', 'months', fallback=12))


        # Worksheet number config
        if self.config.has_option('document', 'worksheet'):
            try:
                sheet_number = int(self.config['document']['worksheet'])
                if sheet_number == 0:
                    sheet_number = 1
                    print('Sheet numbers start at 1, setting worksheet to 1 instead of 0')
            except:
                sheet_number = 1
                print('Invalid worksheet specified in ' + str(self.config_file) + ' defaulting to sheet 1')
        else:
            print('No worksheet specified in ' + str(self.config_file) + ' defaulting to sheet 1')
            sheet_number = 1
        self.ws = load_workbook(work_dir + '/' + work_file).worksheets[sheet_number - 1]


        # Date column config
        if self.config.has_option('document', 'date_column'):
            if str(self.config['document']['date_column']).isalpha():
                self.date_column = self.config['document']['date_column']
            else:
                print('Date column must be a letter')
                return
        else:
            print('No date column defined in ' + str(self.config_file))
            return
        self.cells_start, self.cells_end = self.date_cell_range(self.date_column)
        self.ready = True


    def date_cell_range(self, column):
        """Determines worksheet active rows by first date and last date in input column"""
        last_date_cell = None
        first_date_cell = None
        last_date_counter = 0
        print("Determining column: " + str(column))

        for row in self.ws[column + '1':\
                column + str(len(self.ws[column]))]:
            for cell in row:
                last_date_counter = last_date_counter + 1
                #print(str(cell.value))
                if cell.value is not None:
                    if isinstance(cell.value, datetime.datetime):
                        last_date_cell = last_date_counter
                        if first_date_cell == None:
                            first_date_cell = last_date_counter
        print("Cells are: " + str(first_date_cell) + " " + str(last_date_cell))
        return(str(first_date_cell), str(last_date_cell))


    def column_list(self, column):
        """For readability, returns a list of all cell.value
        for input column, range determined by date_cell_range()"""
        row_cells = []
        for row in self.ws[column + self.cells_start : column + self.cells_end]:
            for cell in row:
                row_cells.append(cell.value)
        return row_cells


    def month_string(self, input_date):
        """returns month as 3 letter string"""
        if isinstance(input_date, datetime.datetime) or isinstance(input_date, datetime.date):
            return input_date.strftime('%b %y')
        else:
            print('Cant convert ' + str(type(input_date)) + ' to month string')
            return None


    def cell_enumerate(self, column):
        """enumerate() but index starts at first cell

        as determined by date_cell_range()"""
        index = int(self.cells_start)
        it_column = iter(column)
        while True:
            try:
                yield (index, it_column.__next__())
            except StopIteration:
                return
            index += 1


    def unique_column_list(self, column):
        """Returns set of column_list()"""
        values = list(set(self.column_list(column)))
        values.sort
        return values


    def diff_month(self, date_1, date_2):
        """Returns how many months apart two dates are"""
        
        #print("year calc: " + str(date_1.year - date_2.year) * 12 )
        #print("month calc: " + str(date_1.month - date_2.month))
        return int((date_1.year - date_2.year) * 12 + date_1.month - date_2.month)


    def curr_month_values(self, column, blanks=False):
        """Totals for input column for this month"""
        this_month = []
        total_occurances = 0
        for index, value in self.cell_enumerate(self.column_list(column)):
            date_cell = self.ws.cell(index, column_index_from_string(self.date_column)).value

            if blanks == False and str(value) == 'None':
                continue #Skips blank cells

            try:
                if date_cell.month == now.month\
                        and date_cell.year == now.year:
                    this_month.append(value)
                    total_occurances += 1
            except AttributeError:
                continue
        return Counter(this_month), total_occurances


    def relative_month_to_string(self, relative_month):
        """Takes relative month as int, returns month string

        int is positive for time in past (think int(x) months ago)"""
        return self.month_string(now
                - dateutil.relativedelta.relativedelta(\
                        months=int(relative_month)))


    def totals_by_month_graph(self, title, months=None):
        """Writes line pgfplot to graph.tex of totals by month"""
        if months == None:
            months = self.months
        column = self.date_column
        months_counter = [0] * months
        symbolic_xcoords = [] # turns into string later
        coordinates = []
        ticks_distance_flag = 0

        for index in range(int(self.cells_end), int(self.cells_start) - 1, -1):
            column_value = self.ws[column + str(index)].value
            try:
                months_ago = self.diff_month(now, column_value)
            except AttributeError:
                #print("attribute error on:" + str(index) + " which is: " + str(type(column_value)))
                continue
            if months_ago > (months - 1):
                #print("neg months on: " + str(months_ago) + " at: " + str(index))
                break
            if months_ago < 0:
                #print("less than months")
                continue
            months_counter[months_ago] += 1

        for i in months_counter:
            if i > 5:
                ticks_distance_flag = 1
                break

        for index, data in reversed(list(enumerate(months_counter))):
            xcoord_name = self.relative_month_to_string(index)
            coordinates.append('(' + xcoord_name + ',' + str(data) + ')')
            symbolic_xcoords.append(xcoord_name + ',')

        symbolic_xcoords = '\n'.join([str(i) for i in symbolic_xcoords])
        coordinates = '\n'.join([str(i) for i in coordinates])

        with open(self.work_dir + '/graph.tex', 'a', encoding='utf-8') as graphs_file:
            graphs_file.write(r'\newpage\addsubsection{' + title + '}')
            graphs_file.write(line_graph_tex[0])
            graphs_file.write(title + ' - ' + fy_date(now))
            graphs_file.write(line_graph_tex[1])
            graphs_file.write(symbolic_xcoords)
            graphs_file.write(line_graph_tex[2])
            if ticks_distance_flag == 0:
                graphs_file.write('ytick distance=1,')
            graphs_file.write(line_graph_tex[3])
            graphs_file.write(coordinates)
            graphs_file.write(line_graph_tex[4])

    def totals_by_month_percent_manual_graph(self, title, column, months=None):
        """Writes percent line pgfplot to graph.tex of totals by month compared to a total from right adjacent column"""
        if months == None:
            months = self.months
        #column = column
        total_column = get_column_letter(column_index_from_string(column) + 1)
        months_counter = [0] * months
        months_total_counter = [0] * months
        symbolic_xcoords = [] # turns into string later
        coordinates = []
        ticks_distance_flag = 0

        for index in range(int(self.cells_end), int(self.cells_start) - 1, -1):
            date_column_value = self.ws[self.date_column + str(index)].value
            compare_column_value = self.ws[column + str(index)].value
            total_column_value = self.ws[total_column + str(index)].value
            try:
                months_ago = self.diff_month(now, date_column_value)
            except AttributeError:
                continue
            if months_ago > (months - 1):
                break
            if months_ago < 0:
                continue
            #months_counter[months_ago] += 1
            if compare_column_value is not None:
                # set months counter directly to newest value found in the column
                # instead of counting instances of date in date column
                months_counter[months_ago] = compare_column_value
            if total_column_value is not None:
                months_total_counter[months_ago] = total_column_value

        for index, count in enumerate(months_counter):
            try:
                months_counter[index] = float(count) / float(months_total_counter[index]) * 100
            except ZeroDivisionError:
                months_counter[index] = 0
                print('Warning: no total specified for ' + self.relative_month_to_string(index))

        for i in months_counter:
            if i > 5:
                ticks_distance_flag = 1
                break

        for index, data in reversed(list(enumerate(months_counter))):
            xcoord_name = self.relative_month_to_string(index)
            coordinates.append('(' + xcoord_name + ',' + format(data, '.2f') + ')')
            symbolic_xcoords.append(xcoord_name + ',')

        symbolic_xcoords = '\n'.join([str(i) for i in symbolic_xcoords])
        coordinates = '\n'.join([str(i) for i in coordinates])

        with open(self.work_dir + '/graph.tex', 'a', encoding='utf-8') as graphs_file:
            graphs_file.write(r'\newpage\addsubsection{' + title + '}')
            graphs_file.write(percent_line_graph_tex[0])
            graphs_file.write(title + ' - ' + fy_date(now))
            graphs_file.write(percent_line_graph_tex[1])
            graphs_file.write(symbolic_xcoords)
            graphs_file.write(percent_line_graph_tex[2])
            if ticks_distance_flag == 0:
                graphs_file.write('ytick distance=1,')
            graphs_file.write(percent_line_graph_tex[3])
            graphs_file.write(coordinates)
            graphs_file.write(percent_line_graph_tex[4])


    def totals_by_month_percent_graph(self, title, months=None):
        """Writes percent line pgfplot to graph.tex of totals by month compared to a total from adjacent column"""
        if months == None:
            months = self.months
        column = self.date_column
        total_column = get_column_letter(column_index_from_string(self.date_column) + 1)
        months_counter = [0] * months
        months_total_counter = [0] * months
        symbolic_xcoords = [] # turns into string later
        coordinates = []
        ticks_distance_flag = 0

        for index in range(int(self.cells_end), int(self.cells_start) - 1, -1):
            column_value = self.ws[column + str(index)].value
            total_column_value = self.ws[total_column + str(index)].value
            try:
                months_ago = self.diff_month(now, column_value)
            except AttributeError:
                continue
            if months_ago > (months - 1):
                break
            if months_ago < 0:
                continue
            months_counter[months_ago] += 1
            if total_column_value is not None:
                months_total_counter[months_ago] = total_column_value

        for i in months_counter:
            if i > 5:
                ticks_distance_flag = 1
                break

        for index, count in enumerate(months_counter):
            try:
                months_counter[index] = float(count) / float(months_total_counter[index]) * 100
            except ZeroDivisionError:
                months_counter[index] = 0
                print('Warning: no total specified for ' + self.relative_month_to_string(index))

        for index, data in reversed(list(enumerate(months_counter))):
            xcoord_name = self.relative_month_to_string(index)
            coordinates.append('(' + xcoord_name + ',' + format(data, '.2f') + ')')
            symbolic_xcoords.append(xcoord_name + ',')

        symbolic_xcoords = '\n'.join([str(i) for i in symbolic_xcoords])
        coordinates = '\n'.join([str(i) for i in coordinates])

        with open(self.work_dir + '/graph.tex', 'a', encoding='utf-8') as graphs_file:
            graphs_file.write(r'\newpage\addsubsection{' + title + '}')
            graphs_file.write(percent_line_graph_tex[0])
            graphs_file.write(title + ' - ' + fy_date(now))
            graphs_file.write(percent_line_graph_tex[1])
            graphs_file.write(symbolic_xcoords)
            graphs_file.write(percent_line_graph_tex[2])
            if ticks_distance_flag == 0:
                graphs_file.write('ytick distance=1,')
            graphs_file.write(percent_line_graph_tex[3])
            graphs_file.write(coordinates)
            graphs_file.write(percent_line_graph_tex[4])


    def totals_by_month_manual_graph(self, title, column, months=None):
        """Writes percent line pgfplot to graph.tex of totals taken directly from column"""
        if months == None:
            months = self.months
        #column = column
        #total_column = get_column_letter(column_index_from_string(column) + 1)
        months_counter = [0] * months
        #months_total_counter = [0] * months
        symbolic_xcoords = [] # turns into string later
        coordinates = []
        ticks_distance_flag = 0

        for index in range(int(self.cells_end), int(self.cells_start) - 1, -1):
            date_column_value = self.ws[self.date_column + str(index)].value
            compare_column_value = self.ws[column + str(index)].value
            #total_column_value = self.ws[total_column + str(index)].value
            try:
                months_ago = self.diff_month(now, date_column_value)
            except AttributeError:
                continue
            if months_ago > (months - 1):
                break
            if months_ago < 0:
                continue
            #months_counter[months_ago] += 1
            if compare_column_value is not None:
                # set months counter directly to newest value found in the column
                # instead of counting instances of date in date column
                months_counter[months_ago] = compare_column_value
            #if total_column_value is not None:
            #    months_total_counter[months_ago] = total_column_value

        #for index, count in enumerate(months_counter):
        #    try:
        #        months_counter[index] = float(count) / float(months_total_counter[index]) * 100
        #    except ZeroDivisionError:
        #        months_counter[index] = 0
        #        print('Warning: no total specified for ' + self.relative_month_to_string(index))

        for i in months_counter:
            if i > 5:
                ticks_distance_flag = 1
                break

        for index, data in reversed(list(enumerate(months_counter))):
            xcoord_name = self.relative_month_to_string(index)
            coordinates.append('(' + xcoord_name + ',' + format(data, '.2f') + ')')
            symbolic_xcoords.append(xcoord_name + ',')

        symbolic_xcoords = '\n'.join([str(i) for i in symbolic_xcoords])
        coordinates = '\n'.join([str(i) for i in coordinates])

        with open(self.work_dir + '/graph.tex', 'a', encoding='utf-8') as graphs_file:
            graphs_file.write(r'\newpage\addsubsection{' + title + '}')
            graphs_file.write(line_graph_tex[0])
            graphs_file.write(title + ' - ' + fy_date(now))
            graphs_file.write(line_graph_tex[1])
            graphs_file.write(symbolic_xcoords)
            graphs_file.write(line_graph_tex[2])
            if ticks_distance_flag == 0:
                graphs_file.write('ytick distance=1,')
            graphs_file.write(line_graph_tex[3])
            graphs_file.write(coordinates)
            graphs_file.write(line_graph_tex[4])


    def values_by_month(self, column, months=None):
        """returns dict of tuples, tuple index is how many months ago data is for"""
        if months == None:
            months = self.months
        catagories_index = {} #dict to get index from string for catagories
        catagories = []       #list of lists for each catagory
        return_dict = {}
        for index, value in reversed(list(self.cell_enumerate(self.column_list(self.date_column)))):
            column_value = str(self.ws[column + str(index)].value) #This must be str to allow sorted()
            try:
                months_ago = self.diff_month(now, value)
            except AttributeError:
                #print("diff_month att error")
                continue
            if months_ago > (months - 1):
                #print("month limit")
                break
            if months_ago < 0:
                #print("neg months")
                continue
            if column_value not in catagories_index:
                catagories_index[column_value] = len(catagories_index)
                catagories.append([0] * months) #index is how many months ago
            catagories[catagories_index.get(column_value)][months_ago] += 1

        for i in catagories_index:
            return_dict[i] = tuple(catagories[catagories_index.get(i)])
        return(return_dict)


    def current_month_graph(self, column, title, blanks=False):
        """Writes bar/pareto pgfplot to graph.tex of current month values"""
        current_data, total_occurances = self.curr_month_values(column, blanks)

        symbolic_xcoords = '\n'.join([str(i) + ',' for i \
                in sorted(current_data, key=current_data.get, reverse=True)])
        coordinates = '\n'.join([str(i).replace("'",'') for i in current_data.most_common()])

        # ytick distance must be set to 1 for values less than 6
        # or the ytick labels will be non whole numbers
        ticks_distance_flag = 0
        occurances = sorted(current_data.values(), reverse=True)

        for i in occurances:
            if i > 5:
                ticks_distance_flag = 1
                break

        pareto = list(itertools.accumulate(occurances))
        pareto = [round(x / total_occurances * 100) for x in pareto]
        pareto = '\n'.join(['(' + str(index) + ',' + str(percent) + ')'\
                for index, percent in enumerate(pareto)])

        with open(self.work_dir + '/graph.tex', 'a', encoding='utf-8') as graphs_file:
            graphs_file.write(r'\newpage\addsubsection{' + title + '}')
            graphs_file.write(bar_graph_tex[0])
            graphs_file.write(title + ' - ' + fy_date(now))
            graphs_file.write(bar_graph_tex[1])
            #TODO make bar width = 0.5/num of bars
            graphs_file.write(symbolic_xcoords)
            graphs_file.write(bar_graph_tex[2])
            if ticks_distance_flag == 0:
                graphs_file.write('ytick distance=1,')
            graphs_file.write(bar_graph_tex[3])
            graphs_file.write(coordinates)
            graphs_file.write(bar_graph_tex[4])
            graphs_file.write(pareto)
            graphs_file.write(bar_graph_tex[5])


    #Uses line graph for one catagory over time
    def monthly_graph(self, column, title, months=None, blanks=False):
        """Writes line pgfplot to graph.tex for one catagory over time"""
        with open(self.work_dir + '/graph.tex', 'a', encoding='utf-8') as graphs_file:
            graphs_file.write(r'\newpage\addsubsection{' + title + ' monthly graphs}\n')

        if months == None:
            months = self.months
        for catagory, month_count_tuple in sorted(self.values_by_month(column, months).items()):

            if blanks == False and str(catagory) == 'None':
                continue #this skips empty cells

            symbolic_xcoords = []
            coordinates = []
            ticks_distance_flag = 0

            # ytick distance must be set to 1 for values less than 6
            # or the ytick labels will be non whole numbers
            for relative_month, occurances in enumerate(month_count_tuple):
                if int(occurances) > 5:
                    ticks_distance_flag = 1
                #print("rel month=" + str(relative_month) + " occ=" + str(occurances))

                xcoord_month = self.relative_month_to_string(relative_month)

                #for use in LaTeX pgfplots package
                symbolic_xcoords.insert(0, xcoord_month + ',')
                coordinates.insert(0, '(' + xcoord_month + ',' + str(occurances) + ')')

            #lists changed to str with newlines to avoid using a loop to append to tex file
            symbolic_xcoords = '\n'.join([str(i) for i in symbolic_xcoords])
            coordinates = '\n'.join([str(i) for i in coordinates])

            with open(self.work_dir + '/graph.tex', 'a', encoding='utf-8') as graphs_file:
                graphs_file.write(r'\addsubsubsection{' + catagory + '}')
                graphs_file.write(line_graph_tex[0])
                graphs_file.write(title + ' - ' + catagory)
                graphs_file.write(line_graph_tex[1])
                graphs_file.write(symbolic_xcoords)
                graphs_file.write(line_graph_tex[2])
                if ticks_distance_flag == 0:
                    graphs_file.write('ytick distance=1,')
                graphs_file.write(line_graph_tex[3])
                graphs_file.write(coordinates)
                graphs_file.write(line_graph_tex[4])

    def custom_page(self, file_path, title, count):
        """Writes a custom page to graph.tex"""
        if os.path.isfile(file_path):
            file_name = os.path.basename(file_path)
            file_ext = file_name.split(".")[ len(file_name.split("."))-1 ]
            new_file_name = str(count) + "." + file_ext
            shutil.copyfile(file_path, os.path.join(self.work_dir, new_file_name))

            with open(self.work_dir + '/graph.tex', 'a', encoding='utf-8') as graphs_file:
                graphs_file.write(r'\newpage\addsubsection{' + title + '}')
                graphs_file.write(custom_page_tex[0])
                graphs_file.write(title)
                graphs_file.write(custom_page_tex[1])
                graphs_file.write(new_file_name)
                graphs_file.write(custom_page_tex[2])
        else:
            print('UserWarning: "' + file_path + '" not found')


    def totals_by_month_percent_manual_graph_with_goal(self, column, title, goal=1.0, avg=3, months=None, legend_labels=None):
        #print("goal percent line")
        column = column.split(',')
        if len(column) == 2:
            #use just two columns because 1st should be date column
            #count_column = column[0]
            cost_column = column[0]
            total_cost_column = column[1]
        else:
            print("totals_by_month_percent_manual_graph_with_goal requires 3 columns")
            for i in column:
                print(i)

        if months == None:
            months = self.months
        column = self.date_column
        months_counter = [0] * months #Count of date occurances in each month
        total_cost_counter = [None] * (months + (avg-1)) #Total cost for each month (must be averaged after)
        cost_counter = [None] * (months + (avg-1)) #WA cost for each month
        symbolic_xcoords = [] # turns into string later
        coordinates = []
        ticks_distance_flag = 0

        #Fill list of WA costs
        for index in range(int(self.cells_end), int(self.cells_start) - 1, -1):
            this_cost_value = self.ws[cost_column + str(index)].value
            date_value = self.ws[column + str(index)].value
            try:
                months_ago = self.diff_month(now, date_value)
            except AttributeError:
                #print("attribute error on:" + str(index) + " which is: " + str(type(this_cost_value)))
                continue
            if months_ago > (months - 1 + (avg-1)): #Extra months for the look back
                #print("neg months on: " + str(months_ago) + " at: " + str(index))
                break
            if months_ago < 0:
                #print("less than months")
                continue
            if cost_counter[months_ago] == None:
                if this_cost_value != None:
                    cost_counter[months_ago] = this_cost_value

        #Fill list of total costs
        for index in range(int(self.cells_end), int(self.cells_start) - 1, -1):
            this_cost_value = self.ws[total_cost_column + str(index)].value
            date_value = self.ws[column + str(index)].value
            try:
                months_ago = self.diff_month(now, date_value)
            except AttributeError:
                #print("attribute error on:" + str(index) + " which is: " + str(type(this_cost_value)))
                continue
            if months_ago > (months - 1 + (avg-1)): #Extra months for the look back
                #print("neg months on: " + str(months_ago) + " at: " + str(index))
                break
            if months_ago < 0:
                #print("less than months")
                continue
            if total_cost_counter[months_ago] == None:
                if this_cost_value != None:
                    total_cost_counter[months_ago] = this_cost_value

        for index in range(len(cost_counter)):
            if cost_counter[index] == None:
                cost_counter[index] = 0.0
        for index in range(len(total_cost_counter)):
            if total_cost_counter[index] == None:
                total_cost_counter[index] = 0.0

        #Find percent with avg look back
        avg_cost_percent = [0.0] * months
        for months_ago in range(months):
            averaged_total_cost = 0.0
            averaged_cost = cost_counter[months_ago]
            for look_back in range(avg):
                averaged_total_cost += total_cost_counter[months_ago + look_back]
                #averaged_cost += cost_counter[months_ago + look_back]
            averaged_total_cost /= avg
            avg_cost_percent[months_ago] = (averaged_cost / averaged_total_cost) * 100.0

        #Fill occurances based on dates entered
        for index in range(int(self.cells_end), int(self.cells_start) - 1, -1):
            column_value = self.ws[column + str(index)].value
            try:
                months_ago = self.diff_month(now, column_value)
            except AttributeError:
                #print("attribute error on:" + str(index) + " which is: " + str(type(column_value)))
                continue
            if months_ago > (months - 1):
                #print("neg months on: " + str(months_ago) + " at: " + str(index))
                break
            if months_ago < 0:
                #print("less than months")
                continue
            months_counter[months_ago] += 1

        for i in months_counter:
            if i > 5:
                ticks_distance_flag = 1
                break

        #Format coordinates and symbolic xcoords for LaTeX
        for index, data in reversed(list(enumerate(months_counter))):
            xcoord_name = self.relative_month_to_string(index)
            coordinates.append('(' + xcoord_name + ',' + str(data) + ')')
            symbolic_xcoords.append(xcoord_name + ',')
        symbolic_xcoords = '\n'.join([str(i) for i in symbolic_xcoords])
        coordinates = '\n'.join([str(i) for i in coordinates])

        cost_coordinates = []
        xcoord_name = 0
        for index, data in reversed(list(enumerate(avg_cost_percent))):
            cost_coordinates.append( '(' + str(xcoord_name) + ',' + str(data) + ')')
            xcoord_name += 1
        cost_coordinates = '\n'.join( [str(i) for i in cost_coordinates] )

        goal_coordinates = []
        goal_coordinates.append( '(0,' + str(goal) + ')' )
        goal_coordinates.append( '(' + str(months-1) + ',' + str(goal) + ')' )
        goal_coordinates = '\n'.join( [str(i) for i in goal_coordinates] )

        #legend_labels = "".join(legend_labels.split())
        legend_labels = legend_labels.split(',')

        with open(self.work_dir + '/graph.tex', 'a', encoding='utf-8') as graphs_file:
            graphs_file.write(r'\newpage\addsubsection{' + title + '}')
            graphs_file.write(cost_avg_graph[0])
            graphs_file.write(title + ' - ' + fy_date(now))
            graphs_file.write(cost_avg_graph[1])
            graphs_file.write(symbolic_xcoords)
            graphs_file.write(cost_avg_graph[2])
            if ticks_distance_flag == 0:
                graphs_file.write('ytick distance=1,')
            graphs_file.write(cost_avg_graph[3])
            graphs_file.write(coordinates)
            graphs_file.write(cost_avg_graph[4])
            graphs_file.write(cost_coordinates)
            graphs_file.write(cost_avg_graph[5])
            graphs_file.write(goal_coordinates)
            graphs_file.write(cost_avg_graph[6])
            if legend_labels is not None:
                if len(legend_labels) == 3:
                    graphs_file.write(legend_labels[0] + ',' + legend_labels[1] + ',' + legend_labels[2])
            else:
                graphs_file.write(title + ',' + "test" + ',' + "goal")
            graphs_file.write(cost_avg_graph[7])

    def month_table(self, column, title, labels, column_widths):
        #print("month table called here")
        column_list = column.split(',')
        num_columns = len(column_list)

        label_list = column_list.copy()
        for index, label in enumerate(labels.split(",")):
            try:
                label_list[index] = label
            except IndexError:
                print("UserWarning: more labels than columns for table: " + column)
        label_config = ''.join(intersperse([r"\textbf{" + x + "}" for x in label_list], " & "))
        #print(label_config)

        width_list = ["Y"] * num_columns
        column_widths = "".join(column_widths.split())
        for index, width in enumerate(column_widths.split(",")):
            try:
                if width.upper() in ("S", "M", "L"):
                    width_list[index] = width.lower()
            except IndexError:
                print("UserWarning: more table widths than columns for table: " + column)
        for i in range(len(width_list)):
            width_list[i] += "@{}"
        width_list = intersperse(width_list, "|", begin="|", end="|")
        tabular_config = ''.join(width_list)
        #print(tabular_config)

        index_list = []
        for index in range(int(self.cells_start), int(self.cells_end)+1):
            date_cell = self.ws.cell(index, column_index_from_string(self.date_column)).value
            try:
                if date_cell.month == now.month\
                        and date_cell.year == now.year:
                    index_list.append(index)
            except AttributeError:
                continue

        if len(index_list) != 0:
            table_data = {}
            for letter in column_list:
                table_data[letter] = list(str(' ') * len(index_list))

            for letter in column_list:
                for loop_index, real_index in enumerate(index_list):
                    this_cell = self.ws.cell(real_index, column_index_from_string(letter)).value
                    #print(this_cell)
                    if this_cell != None:
                        table_data[letter][loop_index] = utf8tolatex(str(this_cell))

            
            # Re-arrange column lists to row lists to print
            table_string = [] #Will be cast to string later
            for index in range(len(index_list)):
                row_list = []
                for letter in column_list:
                    row_list.append(table_data[letter][index])
                table_string.append(''.join(intersperse(row_list, " & ", end=r" \\")))

            table_string = ''.join(intersperse(table_string, "\n\\hline\n", end="\n\\hline\n"))
            #print(table_string)

            with open(self.work_dir + '/graph.tex', 'a', encoding='utf-8') as graphs_file:
                graphs_file.write(r'\newpage\addsubsection{' + title + '}')
                graphs_file.write(latex_table[0])
                graphs_file.write(tabular_config)
                graphs_file.write(latex_table[1])
                graphs_file.write(str(num_columns))
                graphs_file.write(latex_table[2])
                graphs_file.write(title)
                graphs_file.write(latex_table[3])
                graphs_file.write(label_config)
                graphs_file.write(latex_table[4])
                graphs_file.write(str(num_columns))
                graphs_file.write(latex_table[5])
                graphs_file.write(title + "(cont.)")
                graphs_file.write(latex_table[6])
                graphs_file.write(label_config)
                graphs_file.write(latex_table[7])
                graphs_file.write(table_string)
                graphs_file.write(latex_table[8])


    def column_has_data(self, column):
        """Returns true if any data is found in rows of a column"""
        try:
            for r in self.ws[column + '1' : column + str(len(self.ws[column]))]:
                for cell in r:
                    if cell.value is not None:
                        return True
        except:
            return False
        return False


    def compile(self):
        """Calls graph functions based on what is found in config file"""

        if not self.ready:
            return

        print()
        print('\nLoaded config from ' + str(self.config_file) + ':\n')


        with open(self.work_dir + '/graph.tex', 'a', encoding='utf-8') as name_file:
            name_file.write(r'\newpage') #Without newpage the section name in header will be off by one
            name_file.write(r'\addsection{' + self.config.get('document', 'name', fallback=self.doc_name.replace('_', ' ')) + '}')

        # Totals from date column
        if self.config.getboolean('document', 'show_totals', fallback=False):
            self.totals_by_month_graph(title =\
                    self.config.get('document', 'totals_title', fallback=self.doc_name.replace('_', ' ')))

        # Totals of date column compared to a total from adjacent column
        if self.config.getboolean('document', 'show_percent_totals', fallback=False):
            self.totals_by_month_percent_graph(title =\
                    self.config.get('document', 'percent_totals_title', fallback=self.doc_name.replace('_', ' ')))


        for column in self.config.sections():

            if str(column).lower() == 'document':

                #Get list of keys for any custom pages
                custom_list = []
                for setting in list(self.config.items('document')):
                    if 'custom' in setting[0]:
                        if 'title' not in setting[0]:
                            custom_list.append(setting[0])
                custom_list.sort()

                for count, custom in enumerate(custom_list):
                    #print(self.config.get(column, custom))
                    self.custom_page(
                            file_path = self.config.get(column, custom),
                            title = self.config.get(column, custom + "_title", fallback=" "),
                            count = count)

            #Multi column graphs
            elif ',' in column:

                # Current month table
                if self.config.getboolean(column, 'month_table', fallback=False):
                    self.month_table(\
                            title = self.config.get(column, 'table_title', fallback=self.doc_name.replace('_', ' ')),
                            column = column,
                            labels = self.config.get(column, 'labels', fallback=column),
                            column_widths = self.config.get(column, 'column_widths', fallback=" "))

                if self.config.getboolean(column, 'show_count_with_goal', fallback=False):
                    self.totals_by_month_percent_manual_graph_with_goal(\
                            title = self.config.get(column, 'count_with_goal_title', fallback=self.doc_name.replace('_', ' ')),
                            column = column,
                            goal = self.config.get(column, 'goal', fallback=1.0),
                            legend_labels = self.config.get(column, 'legend_labels', fallback=None)
                            )

            #Single column graphs (some graphs are defined by only one column but use data from an adjacent one)
            else:
                if self.column_has_data(column):

                    # Manual percentage line graph
                    if self.config.getboolean(column, 'show_percent_manual_totals', fallback=False):
                        self.totals_by_month_percent_manual_graph(
                                title = self.config.get(column, 'percent_title', fallback=self.doc_name.replace('_', ' ')),
                                column = column)

                    # Manual monthly line graph
                    if self.config.getboolean(column, 'manual_monthly', fallback=False):
                        self.totals_by_month_manual_graph(
                                title = self.config.get(column, 'manual_monthly_title', fallback=self.doc_name.replace('_', ' ')),
                                column = column)

                    # Bar/Pareto graph for current month
                    if self.config.getboolean(column, 'current_month', fallback=False):
                        self.current_month_graph(str(column), str(self.config.get(column, 'title', fallback='')))

                    # Line graph for one catagory over time
                    if self.config.getboolean(column, 'monthly', fallback=False):
                        self.monthly_graph(str(column), str(self.config.get(column, 'title', fallback='')))
                else:
                    print('UserWarning: Column "'
                            + str(column)
                            + '" specified in '
                            + str(self.config_file)
                            + ' does not contain valid data')
        return True
        #This true triggers XeLaTeX to be called

class LogWindow(QtWidgets.QDialog,ui_log.Ui_Dialog):#, UI.MainUI.Ui_MainWindow):
    """Normally hidden log window, also "hosts" the QProcess for calling XeLaTeX"""

    #For use as external tex file to compile with different layouts
    layout_text = {'paper' : r'\usepackage[paperwidth=8.27in, paperheight=11.69in, landscape]{geometry}',
            'screen' : r'\usepackage[paperwidth=7.5in, paperheight=13.33in, landscape]{geometry}'}

    def __init__(self):
        QtWidgets.QDialog.__init__(self)
        self.setupUi(self)
        self.xelatex = QtCore.QProcess(self)
        self.xelatex.setProcessChannelMode(QtCore.QProcess.MergedChannels)
        self.xelatex.readyRead.connect(self.stdout_and_err_Ready)
        self.xelatex.finished.connect(self.done_statement)
        self.xelatex.stateChanged.connect(self.menu_disable)
        self.proc_count = 0 #Counter to run XeLaTeX four times (three for accurate TOC, fourth for other layout)
        self.xelatex_path_config()
        self.buttonBox.button(QtWidgets.QDialogButtonBox.Abort).clicked.connect(self.xelatex.kill)

        self.outputbox_2.setMaximumBlockCount(50)
        self.outputbox.setReadOnly(True)
        self.outputbox_2.setReadOnly(True)
        font = QtGui.QFontDatabase.systemFont(QtGui.QFontDatabase.FixedFont)
        self.outputbox.setFont(font)
        self.outputbox_2.setFont(font)


    def menu_disable(self):
        """Disable menu and set running icon animation while XeLaTeX is running"""
        if self.xelatex.state() == 0:
            trayIcon.menu.setEnabled(True)
            trayIcon.anim_icon.stop()
            trayIcon.setIcon(trayIcon.still_icon)
        else:
            trayIcon.menu.setEnabled(False)
            trayIcon.anim_icon.start()
            trayIcon.setIcon(trayIcon.still_icon)

    def done_statement(self):
        """Show log if XeLaTeX exits with error

        XeLaTeX is called with '--halt-on-error' for this to work"""
        if self.xelatex.exitCode() == 0:
            return
        else:
            print('XeLaTeX Error!')
            self.show()
            self.proc_count = 4
            shutil.rmtree(self.work_dir)

    def xelatex_path_config(self):
        """Set path to XeLaTeX based on what system is running"""
        if os.name == "nt":
            self.xelatex_path = 'texlive/bin/win32/xelatex.exe'
        elif os.name == "posix":
            self.xelatex_path = 'xelatex'

    def compile_tex(self, work_dir, name, layout):
        """Called to run XeLaTeX if Grapher succesfully runs"""

        self.outputbox.clear()
        self.work_dir = work_dir

        def layout_name(work_dir, name, layout):
            """Called after xelatex finishes, manages proc_count, and renames final output pdf"""
            new_name = output_folder + '/' + filename_noext(name) + ' ' + now.strftime('%b%y') + ' ' + layout + '.pdf'

            if self.xelatex.exitCode() == 0:
                # After third and sixth runs, copy the file and print message
                if self.proc_count in (2, 5):
                    shutil.copy(work_dir + '/present.pdf', new_name)
                    print('XeLaTeX: ' + ' ' + now.strftime('%B %Y') + ' "' + layout + '" done')
                    if self.proc_count == 5:
                        shutil.rmtree(work_dir)

                self.proc_count += 1

                # Must be ran three times for each layout
                # Table of contents sometimes takes different amount of space
                # between the layouts
                if self.proc_count in (1, 2):
                    log.compile_tex(work_dir, name, 'screen')
                    log.xelatex.finished.disconnect()
                    log.xelatex.finished.connect(lambda: layout_name(work_dir, name, 'screen'))

                elif self.proc_count in (3, 4, 5):
                    log.compile_tex(work_dir, name, 'paper')
                    log.xelatex.finished.disconnect()
                    log.xelatex.finished.connect(lambda: layout_name(work_dir, name, 'paper'))

                else:
                    self.proc_count = 0
                    if os.name == 'nt':
                        os.startfile('output_folder')

        with open(work_dir + '/layout.tex', 'w', encoding='utf-8') as layout_file:
            layout_file.write(self.layout_text.get(layout))


        self.xelatex.setWorkingDirectory(work_dir)
        self.xelatex.start(self.xelatex_path, ['--halt-on-error', 'present'])
        self.xelatex.finished.disconnect()
        self.xelatex.finished.connect(self.done_statement)
        self.xelatex.finished.connect(lambda: layout_name(work_dir, name, 'screen'))

    def append(self, text):
        """Write stdout from XeLaTeX to outputbox"""
        cursor = self.outputbox.textCursor()
        cursor.movePosition(cursor.End)
        cursor.insertText(text)
        cursor.movePosition(cursor.End)
        self.outputbox.verticalScrollBar().setSliderPosition\
                (self.outputbox.verticalScrollBar().maximum())

    def stdout_and_err_Ready(self):
        """Feed stdout from XeLaTeX to self.append()"""
        text = str(self.xelatex.readAll(), 'utf-8')
        self.append(text)


class SystemTrayIcon(QtWidgets.QSystemTrayIcon):
    """Program "lives" here, treat like mainwindow, call other functions and classes from here"""

    def __init__(self, icon, parent=None):
        QtWidgets.QSystemTrayIcon.__init__(self, icon, parent)

        self.menu = QtWidgets.QMenu(parent)

        self.setContextMenu(self.menu)
        self.activated.connect(self.tray_clicked)
        self.click_timer = QtCore.QTimer(self)
        self.click_timer.setSingleShot(True)
        self.click_timer.timeout.connect(self.single_click)

        #Used in write() for sys.stdout
        self.print_queue = []
        self.time_stamp = ''
        self.time_format_str = '%H:%M:%S: '
        self.rebuild_menu()

        self.still_icon = icon
        self.anim_icon = QtGui.QMovie('resource/spin.gif', parent = self)
        self.anim_icon.frameChanged.connect(self.update_icon)

    def rebuild_menu(self):
        """Rebuilds menu to add an edit option for any ini files found"""
        def edit_base_function(config_file):
            edit_config = EditConfig(None, config_file)
            edit_config.exec_() #Block until closed

        self.menu.clear()

        open_xlsx = self.menu.addAction('Open spreadsheets folder')
        open_xlsx.triggered.connect(lambda: choose_open_folder())
        self.menu.addSeparator()
        for item in os.listdir('resource'):
            if item.endswith('.ini'):
                self.menu.addAction('Edit ' + item).triggered.connect\
                        (functools.partial(edit_base_function, config_file = 'resource' + '/' + item))
        self.menu.addSeparator()
        run_month = self.menu.addAction('Set graph month')
        run_month.triggered.connect(self.get_run_month)
        exitAction = self.menu.addAction("Exit")
        exitAction.triggered.connect(self.hide_and_quit)#QtWidgets.qApp.quit)

    def hide_and_quit(self):
        """hide tray icon before quitting so it doesnt linger in windows taskbar"""
        self.hide()
        QtWidgets.qApp.quit()

    def update_icon(self):
        """called while animated icon (QMovie) is playing"""
        temp_anim_icon = QtGui.QIcon()
        temp_anim_icon.addPixmap(self.anim_icon.currentPixmap())
        self.setIcon(temp_anim_icon)

    def get_run_month(self):
        """Get number as "months ago" for running a graph for the past

        "now" is global"""
        class Prompt(QtWidgets.QDialog, ui_date_set.Ui_Dialog):
            def __init__(self):
                super(Prompt, self).__init__()
                self.setupUi(self)
                self.setModal(True)
                self.spinBox.valueChanged.connect(self.update_string)
                font = QtGui.QFontDatabase.systemFont(QtGui.QFontDatabase.FixedFont)
                font.setPointSize(20)
                self.label.setFont(font)
                self.buttonBox.button(QtWidgets.QDialogButtonBox.Ok).clicked.connect(self.global_now)

            def showEvent(self, event):
                """Make month label and spinbox match when reopening window"""
                super(Prompt, self).showEvent(event)
                self.update_string()

            def update_string(self):
                """match date string to spinbox current value"""
                temp_now = (datetime.date.today() - dateutil.relativedelta.relativedelta(months = self.spinBox.value()))
                temp_now = temp_now.strftime('%B %Y')
                self.label.setText(temp_now)

            def global_now(self, event):
                """set global now variable and print date"""
                global now
                now = (datetime.date.today() - dateutil.relativedelta.relativedelta(months = self.spinBox.value()))
                print('Setting Graph date to: ' + now.strftime('%B %Y'))
        prompt = Prompt()
        prompt.exec_()


    def flush(self, *arg, **args):
        """Used only to suppress errors"""
        pass

    def write(self, text):
        """sys.stdout directed to a system tray message

        system tray message is printed from a queue,
        old lines in the queue are cleared by a QTimer.
        time stamp added to tray message to differentiate between two messages in queue.
        Full stdout with no queue is appended to log window"""
        def append(text):
            """Write to right side text box"""
            cursor = log.outputbox_2.textCursor()
            cursor.movePosition(cursor.End)
            cursor.insertText(text)
            cursor.movePosition(cursor.End)
            log.outputbox.verticalScrollBar().setSliderPosition\
                    (log.outputbox.verticalScrollBar().maximum())
        append(text)
        text = text.rstrip()
        if len(text) == 0:
            return

        #Adds timestamp only if different than previous one
        if datetime.datetime.strftime(datetime.datetime.now(),
                self.time_format_str) == self.time_stamp:
            self.print_queue.append(text)
        else:
            self.print_queue.append(datetime.datetime.strftime(datetime.datetime.now(),
                self.time_format_str) + text)

        self.time_stamp = datetime.datetime.strftime(datetime.datetime.now(), self.time_format_str)
        queue_timer = QtCore.QTimer(self)
        queue_timer.setSingleShot(True)
        queue_timer.timeout.connect(self.clear_queue_line)
        queue_timer.start(4000)
        self.showMessage('PfaudSec spreadsheet:', '\n'.join([str(i) for i in self.print_queue]), icon)

    def clear_queue_line(self):
        """Called from a QTimer to remove old lines from system tray message queue"""
        try:
            del self.print_queue[0]
        except IndexError:
            pass

    def tray_clicked(self, reason):
        """Determine single or double click with a QTimer"""
        if reason == QtWidgets.QSystemTrayIcon.Trigger:
            self.click_timer.start(QtWidgets.qApp.doubleClickInterval())

        elif reason == QtWidgets.QSystemTrayIcon.DoubleClick:
            self.click_timer.stop()
            self.double_click()

    @staticmethod
    def double_click():
        choose_open_folder()

    @staticmethod
    def single_click():
        if log.isVisible():
            log.hide()
        else:
            log.show()

sys.excepthook = except_box
app = QtWidgets.QApplication(sys.argv)
app.setQuitOnLastWindowClosed(False)
app.setStyleSheet(qdarkstyle.load_stylesheet_pyqt5())
output_folder = folder_check('output_folder')


#Attempts to load fonts, if not found will prompt for password
font_decrypt = font_load.Prompt()
if not font_decrypt.check_success():
    try:
        import Cryptodome
    except:
        pass
    if 'Cryptodome' in sys.modules:
        font_decrypt.show()

#This will change if user sets a different month
now = datetime.date.today()

icon = QtGui.QIcon('resource/logo.ico')
trayIcon = SystemTrayIcon(icon)
trayIcon.show()
sys.stdout = trayIcon
log = LogWindow()


def choose_open_folder():
    """Input folder prompt"""
    input_folder = str(QtWidgets.QFileDialog.getExistingDirectory(None, "Select Job Documents Directory"))
    if input_folder:
        work_dir = tempfile.mkdtemp(prefix='PfSpreadsheet_')
        folder_check(work_dir)
        template_dir = 'TeX'

        try:
            shutil.copytree(template_dir + '/font/', work_dir + '/font/')
        except:
            pass

        for item in os.listdir(template_dir):
            if item.endswith('.tex'):
                shutil.copyfile(template_dir + '/' + item, work_dir + '/' + item)

        folder_name = input_folder.split('/')
        folder_name = folder_name[len(folder_name) - 1]
        folder_name = folder_name.replace('_', ' ')
        with open(work_dir + '/name.tex', 'w', encoding='utf-8') as name_file:
            name_file.write(folder_name + r'\\' + '\n' + fy_date(now))

        for input_file in os.listdir(input_folder):
            if input_file.endswith('.xlsx'):
                doc_name = input_file.split('/')
                doc_name = doc_name[len(doc_name) - 1]
                doc_nospace = doc_name.replace(' ', '_')
                try:
                    shutil.copyfile(input_folder + '/' + input_file, work_dir + '/' + doc_nospace)
                except shutil.SameFileError:
                    pass
                #Append pgfplots tex text to graph.tex for each spreadsheet
                Grapher(work_dir, doc_nospace, doc_name).compile()
        log.proc_count = 0
        log.compile_tex(work_dir, folder_name, 'screen')

sys.exit(app.exec_())
